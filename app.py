from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import pdfplumber
from datetime import datetime
import os
import tempfile
from typing import Dict
import logging
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app = FastAPI(title="Bank Reconciliation System", version="3.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Directories
BASE_DIR = "/tmp/bank_reconciliation"
REPORTS_DIR = os.path.join(BASE_DIR, "reports")
METADATA_FILE = os.path.join(BASE_DIR, "reports_metadata.json")

for directory in [BASE_DIR, REPORTS_DIR]:
    os.makedirs(directory, exist_ok=True)


class BankReconciliation:
    """Bank reconciliation with accurate categorization"""
    
    def __init__(self, data_entry_df: pd.DataFrame, bank_statement_path: str):
        # Filter only entries with journal numbers
        self.original_df = data_entry_df.copy()
        self.data_entry_df = data_entry_df[data_entry_df['ChequeDDNo'].notna()].copy()
        # Keep original format - don't convert ChequeDDNo
        # For matching, we'll use string representation internally
        self.data_entry_df['_ChequeDDNo_str'] = self.data_entry_df['ChequeDDNo'].astype(str).str.strip()
        
        self.bank_df = self.parse_bank_pdf(bank_statement_path)
        self.bank_df['journal_number'] = self.bank_df['journal_number'].astype(str).str.strip()
        
        logger.info(f"Data Entry: {len(self.data_entry_df)} valid entries (with journal numbers)")
        logger.info(f"Bank Statement: {len(self.bank_df)} transactions")
    
    def parse_bank_pdf(self, pdf_path: str) -> pd.DataFrame:
        """Parse Bank of Bhutan statement"""
        transactions = []
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    
                    for table in tables:
                        if not table or len(table) < 2:
                            continue
                        
                        header_idx = -1
                        for idx, row in enumerate(table):
                            if any('JOURNAL' in str(cell).upper() for cell in row if cell):
                                header_idx = idx
                                break
                        
                        if header_idx == -1:
                            continue
                        
                        for row in table[header_idx + 1:]:
                            if not row or len(row) < 5:
                                continue
                            
                            if any(kw in str(row).upper() for kw in ['TOTAL', 'OPENING', 'CLOSING', 'STATEMENT', 'BALANCE AS']):
                                continue
                            
                            try:
                                post_date = str(row[0]).strip() if row[0] else ''
                                particulars = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                                journal_number = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                                amount_str = str(row[6]).strip() if len(row) > 6 and row[6] else '0'
                                
                                if not journal_number or journal_number == 'None' or len(journal_number) < 3:
                                    continue
                                
                                amount = self.parse_amount(amount_str)
                                
                                if amount > 0:
                                    transactions.append({
                                        'date': post_date,
                                        'journal_number': journal_number,
                                        'description': particulars,
                                        'amount': amount
                                    })
                            except:
                                continue
            
            logger.info(f"Extracted {len(transactions)} transactions from PDF")
            return pd.DataFrame(transactions)
            
        except Exception as e:
            logger.error(f"PDF parsing error: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to parse bank statement: {str(e)}")
    
    def parse_amount(self, amount_str: str) -> float:
        """Parse amount to float"""
        if pd.isna(amount_str) or not amount_str:
            return 0.0
        
        amount_str = str(amount_str).replace(',', '').replace('Nu.', '').replace('BTN', '').strip()
        
        try:
            return float(amount_str)
        except:
            return 0.0
    
    def reconcile(self) -> Dict:
        """Main reconciliation logic"""
        
        all_entries = []  # For Sheet 1: All Entries
        matched_entries = []  # For Sheet 2
        adjusted_entries = []  # For Sheet 3
        unmatched_entries = []  # For Sheet 4
        unregistered_entries = []  # For Sheet 5
        
        bank_matched = set()
        
        # Process each data entry
        for idx, row in self.data_entry_df.iterrows():
            journal = row['_ChequeDDNo_str']
            amount = float(row['Amount'])
            
            # Find match in bank
            bank_match = self.bank_df[self.bank_df['journal_number'] == journal]
            
            if bank_match.empty:
                # UNMATCHED - Not in bank
                row_dict = row.to_dict()
                row_dict['Status'] = 'Not Found in Bank'
                row_dict['Bank_Amount'] = ''
                row_dict['Difference'] = ''
                all_entries.append(row_dict)
                unmatched_entries.append(row_dict)
                
            else:
                bank_amount = float(bank_match.iloc[0]['amount'])
                bank_matched.add(bank_match.index[0])
                
                if abs(bank_amount - amount) < 1.0:
                    # MATCHED
                    row_dict = row.to_dict()
                    row_dict['Status'] = 'Matched'
                    row_dict['Bank_Amount'] = bank_amount
                    row_dict['Difference'] = 0
                    all_entries.append(row_dict)
                    matched_entries.append(row_dict)
                    
                else:
                    # AMOUNT MISMATCH
                    difference = bank_amount - amount
                    
                    # Original entry
                    row_dict = row.to_dict()
                    row_dict['Status'] = 'Amount Mismatch'
                    row_dict['Bank_Amount'] = bank_amount
                    row_dict['Difference'] = difference
                    all_entries.append(row_dict)
                    adjusted_entries.append(row_dict)
                    
                    # Adjustment entry (new row below)
                    adj_dict = row_dict.copy()
                    adj_dict['Amount'] = difference
                    adj_dict['Status'] = 'Adjustment Entry'
                    adj_dict['Remarks'] = f'Adjustment for Journal {journal}: Bank={bank_amount}, Entry={amount}, Diff={difference}'
                    all_entries.append(adj_dict)
                    adjusted_entries.append(adj_dict)
        
        # UNREGISTERED - Bank entries not in data entry
        unregistered = self.bank_df[~self.bank_df.index.isin(bank_matched)]
        
        for idx, bank_row in unregistered.iterrows():
            unregistered_entries.append({
                'PaymentDate': bank_row.get('date', ''),
                'CustomerName': 'UNREGISTERED',
                'ChequeDDNo': bank_row['journal_number'],
                'Amount': bank_row['amount'],
                'PaymentMode': 'BANK TRANSFER',
                'Bank': 'Bank of Bhutan',
                'Remarks': bank_row.get('description', ''),
                'Status': 'Not in Data Entry',
                'Bank_Amount': bank_row['amount']
            })
        
        return {
            'all_entries': pd.DataFrame(all_entries),
            'matched': pd.DataFrame(matched_entries),
            'adjusted': pd.DataFrame(adjusted_entries),
            'unmatched': pd.DataFrame(unmatched_entries),
            'unregistered': pd.DataFrame(unregistered_entries)
        }
    
    def generate_report(self, output_path: str) -> Dict:
        """Generate comprehensive Excel report"""
        
        logger.info("Starting reconciliation...")
        categories = self.reconcile()
        
        # Calculate totals
        total_entered = self.data_entry_df['Amount'].sum()
        total_bank = self.bank_df['amount'].sum()
        
        adjusted_df = categories['adjusted']
        adjustment_amount = adjusted_df[adjusted_df['Status'] == 'Adjustment Entry']['Amount'].sum() if len(adjusted_df) > 0 else 0
        
        mismatches_count = len(adjusted_df[adjusted_df['Status'] == 'Amount Mismatch']) if len(adjusted_df) > 0 else 0
        
        stats = {
            'total_entries': int(len(self.data_entry_df)),
            'matched': int(len(categories['matched'])),
            'mismatches': int(mismatches_count),
            'unmatched': int(len(categories['unmatched'])),
            'unregistered': int(len(categories['unregistered'])),
            'total_bank_transactions': int(len(self.bank_df)),
            'total_entered_amount': float(total_entered),
            'total_adjustment_amount': float(adjustment_amount),
            'total_after_adjustment': float(total_entered + adjustment_amount),
            'total_bank_amount': float(total_bank),
            'amount_difference': float(total_bank - (total_entered + adjustment_amount))
        }
        
        # Create Excel report
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: All Entries
            if len(categories['all_entries']) > 0:
                categories['all_entries'].to_excel(writer, sheet_name='All Entries', index=False)
            else:
                pd.DataFrame({'Message': ['No entries found']}).to_excel(writer, sheet_name='All Entries', index=False)
            
            # Sheet 2: Matched
            if len(categories['matched']) > 0:
                categories['matched'].to_excel(writer, sheet_name='✓ Matched', index=False)
            else:
                pd.DataFrame({'Message': ['No matched transactions']}).to_excel(writer, sheet_name='✓ Matched', index=False)
            
            # Sheet 3: Adjusted
            if len(categories['adjusted']) > 0:
                categories['adjusted'].to_excel(writer, sheet_name='⚠ Adjusted', index=False)
            else:
                pd.DataFrame({'Message': ['No adjustments needed']}).to_excel(writer, sheet_name='⚠ Adjusted', index=False)
            
            # Sheet 4: Unmatched
            if len(categories['unmatched']) > 0:
                categories['unmatched'].to_excel(writer, sheet_name='✗ Unmatched', index=False)
            else:
                pd.DataFrame({'Message': ['All entries found in bank']}).to_excel(writer, sheet_name='✗ Unmatched', index=False)
            
            # Sheet 5: Unregistered
            if len(categories['unregistered']) > 0:
                categories['unregistered'].to_excel(writer, sheet_name='⚠ Unregistered', index=False)
            else:
                pd.DataFrame({'Message': ['All bank transactions registered']}).to_excel(writer, sheet_name='⚠ Unregistered', index=False)
            
            # Sheet 6: Summary
            summary_data = {
                'Description': [
                    '📊 TOTAL DATA ENTRIES',
                    '',
                    '✓ Matched Transactions',
                    '⚠ Amount Mismatches',
                    '✗ Unmatched (not in bank)',
                    '⚠ Unregistered (not in sheet)',
                    '',
                    '🏦 TOTAL BANK TRANSACTIONS',
                    '',
                    '💰 AMOUNT RECONCILIATION',
                    '  Total Entered Amount',
                    '  Total Adjustment Amount',
                    '  Total After Adjustment',
                    '',
                    '  Total Bank Amount',
                    '',
                    '  DIFFERENCE',
                    '',
                    '📈 RECONCILIATION ACCURACY',
                ],
                'Value': [
                    stats['total_entries'],
                    '',
                    stats['matched'],
                    stats['mismatches'],
                    stats['unmatched'],
                    stats['unregistered'],
                    '',
                    stats['total_bank_transactions'],
                    '',
                    '',
                    f"Nu. {stats['total_entered_amount']:,.2f}",
                    f"Nu. {stats['total_adjustment_amount']:,.2f}",
                    f"Nu. {stats['total_after_adjustment']:,.2f}",
                    '',
                    f"Nu. {stats['total_bank_amount']:,.2f}",
                    '',
                    f"Nu. {stats['amount_difference']:,.2f}",
                    '',
                    f"{(stats['matched']/stats['total_entries']*100):.1f}%" if stats['total_entries'] > 0 else '0%',
                ]
            }
            
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='📊 Summary', index=False)
            
            # Sheet 7: Bank Statement Data (for convenience)
            if len(self.bank_df) > 0:
                bank_display_df = self.bank_df.copy()
                # Rename columns for better readability
                bank_display_df.columns = ['Date', 'Journal #', 'Description', 'Amount']
                bank_display_df.to_excel(writer, sheet_name='🏦 Bank Statement', index=False)
            else:
                pd.DataFrame({'Message': ['No bank data to display']}).to_excel(writer, sheet_name='🏦 Bank Statement', index=False)
        
        # Now add SUM formulas only to Amount column in transaction sheets
        wb = load_workbook(output_path)
        
        # Sheets that should have TOTAL row (only Amount column)
        sheets_with_totals = ['All Entries', '✓ Matched', '⚠ Adjusted', '✗ Unmatched', '⚠ Unregistered', '🏦 Bank Statement']
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Skip Summary sheet - no TOTAL row needed
            if sheet_name == '📊 Summary':
                continue
            
            # Add TOTAL only to Amount column for transaction sheets
            if sheet_name in sheets_with_totals and ws.max_row > 1:
                total_row = ws.max_row + 1
                
                # Find Amount column
                amount_col_idx = None
                for col_idx in range(1, ws.max_column + 1):
                    header_cell = ws.cell(row=1, column=col_idx)
                    if header_cell.value == 'Amount':
                        amount_col_idx = col_idx
                        break
                
                if amount_col_idx:
                    # Add TOTAL label in first column
                    ws.cell(row=total_row, column=1).value = "TOTAL"
                    # Add SUM formula only for Amount column
                    formula = f"=SUM({get_column_letter(amount_col_idx)}2:{get_column_letter(amount_col_idx)}{ws.max_row - 1})"
                    ws.cell(row=total_row, column=amount_col_idx).value = formula
        
        wb.save(output_path)
        
        logger.info(f"Report generated: {output_path}")
        return stats


def save_report_metadata(filename: str, stats: Dict, data_entry_name: str, bank_statement_name: str):
    """Save report metadata"""
    metadata = []
    
    # Try to load existing metadata, but handle corruption gracefully
    if os.path.exists(METADATA_FILE):
        try:
            with open(METADATA_FILE, 'r') as f:
                metadata = json.load(f)
        except (json.JSONDecodeError, IOError) as e:
            logger.warning(f"Metadata file corrupted, starting fresh: {e}")
            metadata = []
    
    metadata.append({
        'filename': filename,
        'timestamp': datetime.now().isoformat(),
        'data_entry_file': data_entry_name,
        'bank_statement_file': bank_statement_name,
        'stats': stats
    })
    
    metadata = metadata[-50:]
    
    with open(METADATA_FILE, 'w') as f:
        json.dump(metadata, f, indent=2)


@app.get("/", response_class=HTMLResponse)
async def home():
    """Serve UI with orange/yellow theme"""
    
    html_content = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Bank Reconciliation Pro</title>
        <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
        <style>
            * {
                margin: 0;
                padding: 0;
                box-sizing: border-box;
            }
            
            :root {
                --primary: #f59e0b;
                --primary-dark: #d97706;
                --secondary: #fbbf24;
                --success: #10b981;
                --warning: #f59e0b;
                --danger: #ef4444;
                --gray-50: #f9fafb;
                --gray-100: #f3f4f6;
                --gray-200: #e5e7eb;
                --gray-600: #4b5563;
                --gray-700: #374151;
                --gray-800: #1f2937;
                --gray-900: #111827;
            }
            
            body {
                font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
                background: linear-gradient(135deg, #f59e0b 0%, #fbbf24 100%);
                min-height: 100vh;
                padding: 2rem;
                color: var(--gray-900);
            }
            
            .container {
                max-width: 1200px;
                margin: 0 auto;
            }
            
            .header {
                text-align: center;
                margin-bottom: 3rem;
                color: white;
                text-shadow: 0 2px 4px rgba(0,0,0,0.1);
            }
            
            .header h1 {
                font-size: 3rem;
                font-weight: 700;
                margin-bottom: 0.5rem;
            }
            
            .header p {
                font-size: 1.25rem;
                opacity: 0.95;
            }
            
            .main-content {
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 2rem;
                margin-bottom: 2rem;
            }
            
            @media (max-width: 968px) {
                .main-content {
                    grid-template-columns: 1fr;
                }
            }
            
            .card {
                background: white;
                border-radius: 16px;
                padding: 2rem;
                box-shadow: 0 10px 40px rgba(0,0,0,0.1);
            }
            
            .upload-card {
                border: 2px dashed var(--gray-300);
                transition: all 0.3s ease;
            }
            
            .upload-card:hover {
                border-color: var(--primary);
                box-shadow: 0 10px 40px rgba(245, 158, 11, 0.2);
            }
            
            .card-header {
                display: flex;
                align-items: center;
                margin-bottom: 1.5rem;
            }
            
            .card-icon {
                width: 48px;
                height: 48px;
                background: var(--primary);
                border-radius: 12px;
                display: flex;
                align-items: center;
                justify-content: center;
                margin-right: 1rem;
                font-size: 1.5rem;
            }
            
            .card-title {
                font-size: 1.25rem;
                font-weight: 600;
                color: var(--gray-800);
            }
            
            .file-input-wrapper {
                position: relative;
            }
            
            .file-input-wrapper input[type=file] {
                position: absolute;
                opacity: 0;
                width: 100%;
                height: 100%;
                cursor: pointer;
            }
            
            .file-input-label {
                display: block;
                padding: 3rem 2rem;
                background: var(--gray-50);
                border: 2px dashed var(--gray-300);
                border-radius: 12px;
                text-align: center;
                cursor: pointer;
                transition: all 0.3s ease;
            }
            
            .file-input-label:hover {
                background: #fef3c7;
                border-color: var(--primary);
            }
            
            .file-input-label.has-file {
                background: #fef3c7;
                border-color: var(--primary);
            }
            
            .file-icon {
                font-size: 3rem;
                margin-bottom: 1rem;
            }
            
            .file-text {
                font-size: 0.875rem;
                color: var(--gray-600);
            }
            
            .file-name {
                margin-top: 1rem;
                padding: 0.75rem 1rem;
                background: white;
                border-radius: 8px;
                font-size: 0.875rem;
                color: var(--gray-700);
                font-weight: 500;
                border: 1px solid var(--gray-200);
            }
            
            .full-width-card {
                grid-column: 1 / -1;
            }
            
            .btn {
                width: 100%;
                padding: 1.25rem 2rem;
                font-size: 1.125rem;
                font-weight: 600;
                color: white;
                background: linear-gradient(135deg, var(--primary) 0%, var(--primary-dark) 100%);
                border: none;
                border-radius: 12px;
                cursor: pointer;
                transition: all 0.3s ease;
                box-shadow: 0 4px 12px rgba(245, 158, 11, 0.3);
            }
            
            .btn:hover:not(:disabled) {
                background: linear-gradient(135deg, var(--primary-dark) 0%, #b45309 100%);
                transform: translateY(-2px);
                box-shadow: 0 6px 20px rgba(245, 158, 11, 0.4);
            }
            
            .btn:disabled {
                opacity: 0.5;
                cursor: not-allowed;
                transform: none;
            }
            
            .status-card {
                display: none;
                margin-top: 2rem;
            }
            
            .status-card.show {
                display: block;
            }
            
            .status-success {
                background: #d1fae5;
                border-left: 4px solid var(--success);
            }
            
            .loading {
                text-align: center;
                padding: 2rem;
                display: none;
            }
            
            .spinner {
                width: 60px;
                height: 60px;
                border: 4px solid var(--gray-200);
                border-top: 4px solid var(--primary);
                border-radius: 50%;
                animation: spin 1s linear infinite;
                margin: 0 auto 1rem;
            }
            
            @keyframes spin {
                0% { transform: rotate(0deg); }
                100% { transform: rotate(360deg); }
            }
            
            .stats-grid {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
                gap: 1rem;
                margin: 1.5rem 0;
            }
            
            .stat-box {
                padding: 1rem;
                background: var(--gray-50);
                border-radius: 8px;
                text-align: center;
            }
            
            .stat-value {
                font-size: 2rem;
                font-weight: 700;
                color: var(--primary);
            }
            
            .stat-label {
                font-size: 0.75rem;
                color: var(--gray-600);
                margin-top: 0.25rem;
            }
            
            .action-buttons {
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 1rem;
                margin-top: 1.5rem;
            }
            
            .btn-secondary {
                background: var(--gray-700);
                color: white;
                padding: 0.875rem 1.5rem;
                border-radius: 8px;
                text-decoration: none;
                text-align: center;
                font-weight: 600;
                transition: all 0.3s ease;
                display: inline-block;
            }
            
            .btn-secondary:hover {
                background: var(--gray-800);
                transform: translateY(-1px);
            }
            
            .reports-section {
                margin-top: 3rem;
            }
            
            .reports-grid {
                display: grid;
                grid-template-columns: repeat(auto-fill, minmax(300px, 1fr));
                gap: 1rem;
                margin-top: 1rem;
            }
            
            .report-card {
                background: white;
                border-radius: 12px;
                padding: 1.5rem;
                box-shadow: 0 2px 8px rgba(0,0,0,0.1);
                transition: all 0.3s ease;
            }
            
            .report-card:hover {
                box-shadow: 0 4px 16px rgba(0,0,0,0.15);
                transform: translateY(-2px);
            }
            
            .report-date {
                font-size: 0.875rem;
                color: var(--gray-600);
                margin-bottom: 1rem;
            }
            
            .report-stats {
                display: grid;
                grid-template-columns: repeat(2, 1fr);
                gap: 0.5rem;
                margin: 1rem 0;
                font-size: 0.875rem;
                color: var(--gray-700);
            }
            
            .report-actions {
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 0.5rem;
                margin-top: 1rem;
            }
            
            .btn-small {
                padding: 0.5rem 1rem;
                font-size: 0.875rem;
                border-radius: 6px;
                border: none;
                cursor: pointer;
                font-weight: 600;
                transition: all 0.3s ease;
            }
            
            .btn-download {
                background: var(--primary);
                color: white;
            }
            
            .btn-download:hover {
                background: var(--primary-dark);
            }
            
            .btn-view {
                background: var(--gray-200);
                color: var(--gray-700);
            }
            
            .btn-view:hover {
                background: var(--gray-300);
            }
            
            .empty-state {
                text-align: center;
                padding: 3rem;
                color: var(--gray-500);
            }

            .amount-display {
                background: #fef3c7;
                padding: 1rem;
                border-radius: 8px;
                margin: 1rem 0;
                border-left: 4px solid var(--primary);
            }

            .amount-label {
                font-size: 0.875rem;
                color: var(--gray-600);
                margin-bottom: 0.25rem;
            }

            .amount-value {
                font-size: 1.5rem;
                font-weight: 700;
                color: var(--primary-dark);
            }

            .file-details {
                background: #f0f9ff;
                border: 1px solid #bfdbfe;
                border-radius: 8px;
                padding: 1rem;
                margin-top: 0.75rem;
                font-size: 0.875rem;
                color: var(--gray-700);
            }

            .detail-item {
                display: flex;
                justify-content: space-between;
                padding: 0.35rem 0;
            }

            .detail-label {
                font-weight: 500;
                color: var(--gray-600);
            }

            .detail-value {
                color: var(--primary-dark);
                font-weight: 600;
            }

            .loading-spinner {
                display: inline-block;
                width: 14px;
                height: 14px;
                border: 2px solid #bfdbfe;
                border-top: 2px solid var(--primary);
                border-radius: 50%;
                animation: spin 0.8s linear infinite;
                margin-right: 0.5rem;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>🏦 Bank Reconciliation Pro</h1>
                <p>Transaction matching & amount reconciliation</p>
            </div>
            
            <div class="main-content">
                <div class="card upload-card">
                    <div class="card-header">
                        <div class="card-icon">📊</div>
                        <div class="card-title">Data Entry File</div>
                    </div>
                    <div class="file-input-wrapper">
                        <input type="file" id="dataEntryFile" accept=".xlsx,.xls" />
                        <label for="dataEntryFile" class="file-input-label" id="dataEntryLabel">
                            <div class="file-icon">📄</div>
                            <div class="file-text">
                                <strong>Click to upload Excel file</strong><br>
                                Must contain ChequeDDNo column
                            </div>
                        </label>
                    </div>
                    <div class="file-name" id="dataFileName" style="display:none;"></div>
                    <div class="file-details" id="dataFileDetails" style="display:none;"></div>
                </div>
                
                <div class="card upload-card">
                    <div class="card-header">
                        <div class="card-icon">🏦</div>
                        <div class="card-title">Bank Statement</div>
                    </div>
                    <div class="file-input-wrapper">
                        <input type="file" id="bankStatementFile" accept=".pdf" />
                        <label for="bankStatementFile" class="file-input-label" id="bankStatementLabel">
                            <div class="file-icon">📑</div>
                            <div class="file-text">
                                <strong>Click to upload PDF file</strong><br>
                                Bank of Bhutan statement format
                            </div>
                        </label>
                    </div>
                    <div class="file-name" id="bankFileName" style="display:none;"></div>
                    <div class="file-details" id="bankFileDetails" style="display:none;"></div>
                </div>
                
                <div class="card full-width-card">
                    <button class="btn" id="reconcileBtn" onclick="reconcile()" disabled>
                        🚀 Start Reconciliation
                    </button>
                    
                    <div class="loading" id="loading">
                        <div class="spinner"></div>
                        <p><strong>Processing your files...</strong></p>
                        <p style="font-size: 0.875rem; color: var(--gray-600); margin-top: 0.5rem;">
                            Reconciling transactions and calculating amounts
                        </p>
                    </div>
                    
                    <div class="status-card card" id="statusCard">
                        <h3 style="margin-bottom: 1rem;">✅ Reconciliation Complete!</h3>
                        
                        <div class="stats-grid" id="statsGrid"></div>
                        
                        <div id="amountDisplay"></div>
                        
                        <div class="action-buttons">
                            <a href="#" id="downloadLink" class="btn-secondary">📥 Download Report</a>
                            <button onclick="viewReports()" class="btn-secondary">📂 View All Reports</button>
                        </div>
                    </div>
                </div>
            </div>
            
            <div class="reports-section" id="reportsSection">
                <div class="card">
                    <h2 style="margin-bottom: 1.5rem;">📂 Recent Reconciliation Reports</h2>
                    <div class="reports-grid" id="reportsGrid">
                        <div class="empty-state">
                            <p>No reports generated yet.<br>Upload files to create your first report.</p>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        
        <script>
            let dataEntryFile = null;
            let bankStatementFile = null;
            
            document.getElementById('dataEntryFile').addEventListener('change', function(e) {
                dataEntryFile = e.target.files[0];
                const label = document.getElementById('dataEntryLabel');
                const fileName = document.getElementById('dataFileName');
                const fileDetails = document.getElementById('dataFileDetails');
                
                if (dataEntryFile) {
                    label.classList.add('has-file');
                    fileName.style.display = 'block';
                    fileName.textContent = `📊 ${dataEntryFile.name}`;
                    
                    // Show loading spinner
                    fileDetails.style.display = 'block';
                    fileDetails.innerHTML = '<span class="loading-spinner"></span> Reading file...';
                    
                    // Fetch file details
                    fetchFileDetails(dataEntryFile, 'excel', fileDetails);
                }
                checkFiles();
            });
            
            document.getElementById('bankStatementFile').addEventListener('change', function(e) {
                bankStatementFile = e.target.files[0];
                const label = document.getElementById('bankStatementLabel');
                const fileName = document.getElementById('bankFileName');
                const fileDetails = document.getElementById('bankFileDetails');
                
                if (bankStatementFile) {
                    label.classList.add('has-file');
                    fileName.style.display = 'block';
                    fileName.textContent = `🏦 ${bankStatementFile.name}`;
                    
                    // Show loading spinner
                    fileDetails.style.display = 'block';
                    fileDetails.innerHTML = '<span class="loading-spinner"></span> Reading file...';
                    
                    // Fetch file details
                    fetchFileDetails(bankStatementFile, 'pdf', fileDetails);
                }
                checkFiles();
            });
            
            function checkFiles() {
                const btn = document.getElementById('reconcileBtn');
                btn.disabled = !(dataEntryFile && bankStatementFile);
            }
            
            async function fetchFileDetails(file, fileType, detailsDiv) {
                const formData = new FormData();
                formData.append('file', file);
                formData.append('file_type', fileType);
                
                try {
                    const response = await fetch('/analyze-file', {
                        method: 'POST',
                        body: formData
                    });
                    
                    if (response.ok) {
                        const details = await response.json();
                        displayFileDetails(details, detailsDiv);
                    } else {
                        detailsDiv.innerHTML = '⚠️ Could not read file details';
                    }
                } catch (error) {
                    detailsDiv.innerHTML = '⚠️ Error reading file: ' + error.message;
                }
            }
            
            function displayFileDetails(details, detailsDiv) {
                let html = '';
                
                if (details.file_type === 'excel') {
                    html = `
                        <div class="detail-item">
                            <span class="detail-label">✅ Status:</span>
                            <span class="detail-value">${details.status}</span>
                        </div>
                        <div class="detail-item">
                            <span class="detail-label">📊 Total Rows:</span>
                            <span class="detail-value">${details.total_rows}</span>
                        </div>
                        <div class="detail-item">
                            <span class="detail-label">🔢 Columns:</span>
                            <span class="detail-value">${details.columns.join(', ')}</span>
                        </div>
                        ${details.has_cheque_column ? 
                            `<div class="detail-item">
                                <span class="detail-label">✅ ChequeDDNo:</span>
                                <span class="detail-value">Found (${details.valid_entries} entries)</span>
                            </div>` : 
                            `<div class="detail-item">
                                <span class="detail-label">⚠️ ChequeDDNo:</span>
                                <span class="detail-value">Not found</span>
                            </div>`
                        }
                        ${details.has_amount_column ? 
                            `<div class="detail-item">
                                <span class="detail-label">💰 Amount Column:</span>
                                <span class="detail-value">Found</span>
                            </div>` : 
                            `<div class="detail-item">
                                <span class="detail-label">⚠️ Amount:</span>
                                <span class="detail-value">Not found</span>
                            </div>`
                        }
                        <div class="detail-item">
                            <span class="detail-label">📄 File Size:</span>
                            <span class="detail-value">${details.file_size}</span>
                        </div>
                    `;
                } else if (details.file_type === 'pdf') {
                    html = `
                        <div class="detail-item">
                            <span class="detail-label">✅ Status:</span>
                            <span class="detail-value">${details.status}</span>
                        </div>
                        <div class="detail-item">
                            <span class="detail-label">📄 Pages:</span>
                            <span class="detail-value">${details.total_pages}</span>
                        </div>
                        <div class="detail-item">
                            <span class="detail-label">📋 Tables Found:</span>
                            <span class="detail-value">${details.tables_found}</span>
                        </div>
                        <div class="detail-item">
                            <span class="detail-label">🔍 Transactions:</span>
                            <span class="detail-value">${details.transaction_count}</span>
                        </div>
                        <div class="detail-item">
                            <span class="detail-label">📄 File Size:</span>
                            <span class="detail-value">${details.file_size}</span>
                        </div>
                    `;
                }
                
                detailsDiv.innerHTML = html;
            }
            
            async function reconcile() {
                const statusCard = document.getElementById('statusCard');
                const loading = document.getElementById('loading');
                const btn = document.getElementById('reconcileBtn');
                
                statusCard.classList.remove('show');
                loading.style.display = 'block';
                btn.disabled = true;
                
                const formData = new FormData();
                formData.append('data_entry', dataEntryFile);
                formData.append('bank_statement', bankStatementFile);
                
                try {
                    const response = await fetch('/reconcile', {
                        method: 'POST',
                        body: formData
                    });
                    
                    const result = await response.json();
                    
                    loading.style.display = 'none';
                    
                    if (response.ok) {
                        displayResults(result);
                        loadReports();
                    } else {
                        alert('Error: ' + (result.detail || 'Reconciliation failed'));
                    }
                } catch (error) {
                    loading.style.display = 'none';
                    alert('Error: ' + error.message);
                }
                
                btn.disabled = false;
            }
            
            function displayResults(result) {
                const statusCard = document.getElementById('statusCard');
                const statsGrid = document.getElementById('statsGrid');
                const amountDisplay = document.getElementById('amountDisplay');
                const downloadLink = document.getElementById('downloadLink');
                
                const stats = result.stats;
                
                statsGrid.innerHTML = `
                    <div class="stat-box">
                        <div class="stat-value">${stats.total_entries}</div>
                        <div class="stat-label">Total Entries</div>
                    </div>
                    <div class="stat-box">
                        <div class="stat-value" style="color: var(--success);">${stats.matched}</div>
                        <div class="stat-label">Matched</div>
                    </div>
                    <div class="stat-box">
                        <div class="stat-value" style="color: var(--warning);">${stats.mismatches}</div>
                        <div class="stat-label">Mismatches</div>
                    </div>
                    <div class="stat-box">
                        <div class="stat-value" style="color: var(--danger);">${stats.unmatched}</div>
                        <div class="stat-label">Unmatched</div>
                    </div>
                    <div class="stat-box">
                        <div class="stat-value" style="color: var(--warning);">${stats.unregistered}</div>
                        <div class="stat-label">Unregistered</div>
                    </div>
                    <div class="stat-box">
                        <div class="stat-value">${stats.total_bank_transactions}</div>
                        <div class="stat-label">Bank Transactions</div>
                    </div>
                `;
                
                amountDisplay.innerHTML = `
                    <div class="amount-display">
                        <div class="amount-label">Total Entered Amount</div>
                        <div class="amount-value">Nu. ${stats.total_entered_amount.toLocaleString('en-IN', {minimumFractionDigits: 2, maximumFractionDigits: 2})}</div>
                    </div>
                    <div class="amount-display">
                        <div class="amount-label">Total Adjustment Amount</div>
                        <div class="amount-value">Nu. ${stats.total_adjustment_amount.toLocaleString('en-IN', {minimumFractionDigits: 2, maximumFractionDigits: 2})}</div>
                    </div>
                    <div class="amount-display">
                        <div class="amount-label">Total After Adjustment</div>
                        <div class="amount-value">Nu. ${stats.total_after_adjustment.toLocaleString('en-IN', {minimumFractionDigits: 2, maximumFractionDigits: 2})}</div>
                    </div>
                    <div class="amount-display">
                        <div class="amount-label">Total Bank Amount</div>
                        <div class="amount-value">Nu. ${stats.total_bank_amount.toLocaleString('en-IN', {minimumFractionDigits: 2, maximumFractionDigits: 2})}</div>
                    </div>
                    <div class="amount-display" style="background: ${Math.abs(stats.amount_difference) < 1 ? '#d1fae5' : '#fee2e2'}; border-color: ${Math.abs(stats.amount_difference) < 1 ? '#10b981' : '#ef4444'};">
                        <div class="amount-label">Difference</div>
                        <div class="amount-value" style="color: ${Math.abs(stats.amount_difference) < 1 ? '#10b981' : '#ef4444'};">Nu. ${stats.amount_difference.toLocaleString('en-IN', {minimumFractionDigits: 2, maximumFractionDigits: 2})}</div>
                    </div>
                `;
                
                downloadLink.href = `/download/${result.filename}`;
                
                statusCard.classList.add('show', 'status-success');
            }
            
            async function loadReports() {
                try {
                    const response = await fetch('/reports');
                    const reports = await response.json();
                    
                    const reportsGrid = document.getElementById('reportsGrid');
                    
                    if (reports.length === 0) {
                        reportsGrid.innerHTML = '<div class="empty-state"><p>No reports generated yet.</p></div>';
                        return;
                    }
                    
                    reportsGrid.innerHTML = reports.map(report => `
                        <div class="report-card">
                            <div class="report-date">📊 ${new Date(report.timestamp).toLocaleString()}</div>
                            <div class="report-stats">
                                <div>✓ Matched: ${report.stats.matched}</div>
                                <div>⚠ Mismatches: ${report.stats.mismatches}</div>
                                <div>✗ Unmatched: ${report.stats.unmatched}</div>
                                <div>⚠ Unregistered: ${report.stats.unregistered}</div>
                            </div>
                            <div style="font-size: 0.875rem; margin-top: 0.5rem; color: var(--gray-600);">
                                Amount: Nu. ${report.stats.total_entered_amount.toLocaleString('en-IN', {minimumFractionDigits: 2})}
                            </div>
                            <div class="report-actions">
                                <a href="/download/${report.filename}" class="btn-small btn-download">Download</a>
                                <a href="/view/${report.filename}" target="_blank" class="btn-small btn-view">View</a>
                            </div>
                        </div>
                    `).join('');
                    
                } catch (error) {
                    console.error('Error loading reports:', error);
                }
            }

            function viewReports() {
                document.getElementById('reportsSection').scrollIntoView({ behavior: 'smooth' });
            }
            
            loadReports();
        </script>
    </body>
    </html>
    """
    
    return HTMLResponse(content=html_content)


@app.post("/analyze-file")
async def analyze_file(
    file: UploadFile = File(...),
    file_type: str = Form("excel")
):
    """Analyze uploaded file and return details"""
    try:
        with tempfile.NamedTemporaryFile(delete=False) as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_path = tmp_file.name
        
        file_size = len(content)
        file_size_mb = file_size / (1024 * 1024)
        file_size_str = f"{file_size_mb:.2f} MB" if file_size_mb > 1 else f"{file_size / 1024:.2f} KB"
        
        if file_type == "excel":
            try:
                df = pd.read_excel(tmp_path, sheet_name=0)
                total_rows = len(df)
                columns = df.columns.tolist()
                
                has_cheque = 'ChequeDDNo' in columns
                has_amount = 'Amount' in columns
                
                # Count valid entries (with ChequeDDNo)
                valid_entries = 0
                if has_cheque:
                    valid_entries = df[df['ChequeDDNo'].notna()].shape[0]
                
                os.unlink(tmp_path)
                
                return {
                    "file_type": "excel",
                    "status": "✅ File read successfully",
                    "total_rows": int(total_rows),
                    "columns": columns,
                    "has_cheque_column": has_cheque,
                    "has_amount_column": has_amount,
                    "valid_entries": int(valid_entries),
                    "file_size": file_size_str
                }
            except Exception as e:
                try:
                    os.unlink(tmp_path)
                except:
                    pass
                logger.error(f"Error reading Excel: {e}")
                return {
                    "file_type": "excel",
                    "status": f"❌ Error: {str(e)[:80]}",
                    "total_rows": 0,
                    "columns": [],
                    "has_cheque_column": False,
                    "has_amount_column": False,
                    "valid_entries": 0,
                    "file_size": file_size_str
                }
        
        elif file_type == "pdf":
            try:
                transaction_count = 0
                total_pages = 0
                
                with pdfplumber.open(tmp_path) as pdf:
                    total_pages = len(pdf.pages)
                    
                    for page in pdf.pages:
                        tables = page.extract_tables()
                        if tables:
                            for table in tables:
                                if table and len(table) > 1:
                                    # Count data rows excluding header row
                                    # Filter out total/summary rows
                                    filtered_rows = sum(1 for row in table[1:] if not any(kw in str(row).upper() for kw in ['TOTAL', 'OPENING', 'CLOSING', 'STATEMENT', 'BALANCE AS']))
                                    transaction_count += filtered_rows
                
                os.unlink(tmp_path)
                
                return {
                    "file_type": "pdf",
                    "status": "✅ File read successfully",
                    "total_pages": int(total_pages),
                    "tables_found": 1 if transaction_count > 0 else 0,
                    "transaction_count": int(transaction_count),
                    "file_size": file_size_str
                }
            except Exception as e:
                try:
                    os.unlink(tmp_path)
                except:
                    pass
                logger.error(f"Error reading PDF: {e}")
                return {
                    "file_type": "pdf",
                    "status": f"❌ Error reading file: {str(e)}",
                    "total_pages": 0,
                    "tables_found": 0,
                    "transaction_count": 0,
                    "file_size": file_size_str
                }
        
        os.unlink(tmp_path)
        return {"error": "Unknown file type"}
    
    except Exception as e:
        logger.error(f"Analyze file error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/reconcile")
async def reconcile_files(
    data_entry: UploadFile = File(...),
    bank_statement: UploadFile = File(...)
):
    """Process reconciliation"""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_excel:
            tmp_excel.write(await data_entry.read())
            excel_path = tmp_excel.name
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_pdf:
            tmp_pdf.write(await bank_statement.read())
            pdf_path = tmp_pdf.name
        
        df = pd.read_excel(excel_path)
        reconciler = BankReconciliation(df, pdf_path)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"Reconciliation_{timestamp}.xlsx"
        output_path = os.path.join(REPORTS_DIR, output_filename)
        
        stats = reconciler.generate_report(output_path)
        
        save_report_metadata(output_filename, stats, data_entry.filename, bank_statement.filename)
        
        os.unlink(excel_path)
        os.unlink(pdf_path)
        
        return {
            "success": True,
            "filename": output_filename,
            "stats": stats
        }
        
    except Exception as e:
        logger.error(f"Reconciliation error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/reports")
async def get_reports():
    """Get list of reports"""
    if not os.path.exists(METADATA_FILE):
        return []
    
    try:
        with open(METADATA_FILE, 'r') as f:
            metadata = json.load(f)
        return list(reversed(metadata))
    except json.JSONDecodeError:
        # If metadata is corrupted, return empty and rebuild it
        logger.warning("Metadata JSON corrupted, starting fresh")
        return []


@app.get("/download/{filename}")
async def download_file(filename: str):
    """Download report"""
    file_path = os.path.join(REPORTS_DIR, filename)
    
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(
        file_path,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename=filename
    )


@app.get("/view/{filename}")
async def view_file(filename: str):
    """View file inline"""
    file_path = os.path.join(REPORTS_DIR, filename)
    
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(
        file_path,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename=filename,
        headers={"Content-Disposition": f"inline; filename={filename}"}
    )


@app.get("/health")
async def health_check():
    return {"status": "healthy", "version": "3.0.0"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
